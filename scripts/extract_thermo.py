import os
import sys
import re
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def write_dict_to_excel(writer, data, sheet_name):
    wb = writer.book
    ws = wb.create_sheet(title=sheet_name)

    start_row = 1

    for key, value in data.items():
        ws.cell(row=start_row, column=1, value=key)
        
        if isinstance(value, list):
            df = pd.DataFrame(value, columns=[key]).T if len(value) <= 3 else pd.DataFrame(value, columns=[key])
        else:
            df = pd.DataFrame([value], columns=[key])
        
        df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, startcol=0, index=False, header=False)
        
        # for row in ws.iter_rows(min_row=start_row+1, max_row=start_row+len(df), min_col=1, max_col=len(df.columns)):
        #     for cell in row:
        #         cell.border = cell.border + openpyxl.styles.Border(top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'), left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'))
        

        start_row += len(df) + 2 

def extract_other(file_path):
    with open(file_path, 'r') as file:
        lines = file.readlines()

    data = {
        'Energies SCF [ Hartree ]': [],
        'Temperature [ K ]': None,
        'Pressure [ atm ]': None,
        'Principal Moments of Inertia [ amu Å^2 ]': [],
        'Molecular Mass [ amu ]': None,
        'Rotational Symmetry Number': None,
        'Rotational Temperatures [ K ]': [],
        'Rotational Constants [ GHZ ]': [],
        'Zero-point Vibrational Energy [ J/mol ]': None,
        'Zero-point Vibrational Energy [ Kcal/mol ]': None,
        'Vibrational Temperatures [ K ]': [],
        'Zero-point Correction [ Hartree/Particle ]': None,
        'Thermal Correction to Energy [ Hartree/Particle ]': None,
        'Thermal Correction to Enthalpy [ Hartree/Particle ]': None,
        'Thermal Correction to Gibbs Free Energy [ Hartree/Particle ]': None,
        'Harmonic Frequencies [ cm^-1 ]': [],
        'Reduced Masses [ amu ]': [],
        'Force Constants [ mDyne/A ]': [],
        'IR Intensities [ km/mol ]': [],
    }

    energy_scf_pattern = re.compile(r'SCF Done:.*?E\(\w+\)\s*=\s*(-?\d+\.\d+)')
    temp_pressure_pattern = re.compile(r'Temperature\s+([\d\.]+)\s+Kelvin.\s+Pressure\s+([\d\.]+)\s+Atm.')
    inertia_pattern = re.compile(r'\s+Eigenvalues --\s+([\d\.]+)\s+([\d\.]+)\s+([\d\.]+)')
    mass_pattern = re.compile(r'Molecular mass:\s+([\d\.]+)\s+amu.')
    symmetry_pattern = re.compile(r'Rotational symmetry number\s+(\d+)')
    rot_temp_pattern = re.compile(r'Rotational temperatures \(Kelvin\)\s+([\d\.\s]+)')
    rot_const_pattern = re.compile(r'Rotational constants \(GHZ\):\s+([\d\.\s]+)')
    zpe_joules_pattern = re.compile(r'Zero-point vibrational energy\s+(-?[\d\.]+) \(Joules/Mol\)')
    zpe_kcal_pattern = re.compile(r'\s+(-?[\d\.]+) \(Kcal/Mol\)')
    vib_temp_pattern1 = re.compile(r'Vibrational temperatures:\s+([\d\.\s]+)')
    zpc_pattern = re.compile(r'Zero-point correction=\s+(-?[\d\.]+) \(Hartree/Particle\)')
    thermal_energy_pattern = re.compile(r'Thermal correction to Energy=\s+(-?[\d\.]+)')
    thermal_enthalpy_pattern = re.compile(r'Thermal correction to Enthalpy=\s+(-?[\d\.]+)')
    thermal_gibbs_pattern = re.compile(r'Thermal correction to Gibbs Free Energy=\s+(-?[\d\.]+)')
    vib_pattern = re.compile(r'Frequencies --\s+(-?[\d\.\s]+)\s+Red\. masses --\s+([\d\.\s]+)\s+Frc consts  --\s+([\d\.\s]+)\s+IR Inten    --\s+([\d\.\s]+)')

    vib_matches = vib_pattern.findall(''.join(lines))

    for match in vib_matches:
        data['Harmonic Frequencies [ cm^-1 ]'].extend(list(map(float, match[0].split())))
        data['Reduced Masses [ amu ]'].extend(list(map(float, match[1].split())))
        data['Force Constants [ mDyne/A ]'].extend(list(map(float, match[2].split())))
        data['IR Intensities [ km/mol ]'].extend(list(map(float, match[3].split())))


    for line in lines:
        energy_scf_match = energy_scf_pattern.search(line)
        temp_pressure_match = temp_pressure_pattern.search(line)
        mass_match = mass_pattern.search(line)
        inertia_match = inertia_pattern.search(line)
        symmetry_match = symmetry_pattern.search(line)
        rot_temp_match = rot_temp_pattern.search(line)
        rot_const_match = rot_const_pattern.search(line)
        zpe_joules_match = zpe_joules_pattern.search(line)
        zpe_kcal_match = zpe_kcal_pattern.search(line)
        vib_temp_match1 = vib_temp_pattern1.search(line)
        zpc_match = zpc_pattern.search(line)
        thermal_energy_match = thermal_energy_pattern.search(line)
        thermal_enthalpy_match = thermal_enthalpy_pattern.search(line)
        thermal_gibbs_match = thermal_gibbs_pattern.search(line)

        if energy_scf_match:
            data['Energies SCF [ Hartree ]'].append(float(energy_scf_match.group(1)))


        if temp_pressure_match:
            data['Temperature [ K ]'] = float(temp_pressure_match.group(1))
            data['Pressure [ atm ]'] = float(temp_pressure_match.group(2))
        
        if inertia_match:
            data['Principal Moments of Inertia [ amu Å^2 ]'] = list(map(float, inertia_match.groups()))

        if mass_match:
            data['Molecular Mass [ amu ]'] = float(mass_match.group(1))

        if symmetry_match:
            data['Rotational Symmetry Number'] = int(symmetry_match.group(1))

        if rot_temp_match:
            data['Rotational Temperatures [ K ]'] = list(map(float, rot_temp_match.group(1).split()))

        if rot_const_match:
            data['Rotational Constants [ GHZ ]'] = list(map(float, rot_const_match.group(1).split()))

        if zpe_joules_match:
            data['Zero-point Vibrational Energy [ J/mol ]'] = float(zpe_joules_match.group(1))

        if zpe_kcal_match:
            data['Zero-point Vibrational Energy [ Kcal/mol ]'] = float(zpe_kcal_match.group(1))

        if vib_temp_match1:
            data['Vibrational Temperatures [ K ]'] = list(map(float, vib_temp_match1.group(1).split()))
            # Copy also the lines below until the first empty line, 
            for vib_line in lines[lines.index(line)+1:]:
                # if wib_line start with "(Kelvin)", get rid of it
                if vib_line.strip().startswith('(Kelvin)'):
                    vib_line = vib_line.replace('(Kelvin)', '')
                if vib_line.strip() == '':
                    break
                data['Vibrational Temperatures [ K ]'].extend(list(map(float, vib_line.split())))

        if zpc_match:
            data['Zero-point Correction [ Hartree/Particle ]'] = float(zpc_match.group(1))

        if thermal_energy_match:
            data['Thermal Correction to Energy [ Hartree/Particle ]'] = float(thermal_energy_match.group(1))

        if thermal_enthalpy_match:
            data['Thermal Correction to Enthalpy [ Hartree/Particle ]'] = float(thermal_enthalpy_match.group(1))

        if thermal_gibbs_match:
            data['Thermal Correction to Gibbs Free Energy [ Hartree/Particle ]'] = float(thermal_gibbs_match.group(1))

    return data

def extract_data_from_log(file_path):
    with open(file_path, 'r') as file:
        lines = file.readlines()

    data = {
        'E': {},
        'Q': {},
        'Other': {},
    }


    e_pattern = re.compile(r'^\s*(Total|Electronic|Translational|Rotational|Vibrational)\s+([\d\.\-]+)\s+([\d\.\-]+)\s+([\d\.\-]+)')
    q_pattern = re.compile(r'^\s*(Total Bot|Total V=0|Vib \(Bot\)|Vib \(V=0\)|Electronic|Translational|Rotational)\s+([\d\.\-D\+]+)\s+([\d\.\-]+)\s+([\d\.\-]+)')


    for line in lines:
        e_match = e_pattern.match(line)
        q_match = q_pattern.match(line)

        if e_match:
            key = e_match.group(1)
            data['E'][key] = {
                'E (Thermal) [KCal/Mol]': float(e_match.group(2)),
                'CV [Cal/Mol-Kelvin]': float(e_match.group(3)),
                'S [Cal/Mol-Kelvin]': float(e_match.group(4))
            }

        if q_match:
            key = q_match.group(1)
            data['Q'][key] = {
                'Q': float(q_match.group(2).replace('D', 'E')),
                'Log10(Q)': float(q_match.group(3)),
                'Ln(Q)': float(q_match.group(4))
            }

    data['Other'] = extract_other(file_path)

    return data

def main(log_file):
    if not log_file.endswith('.log'):
        print("Il file deve avere estensione .log")
        return

    data = extract_data_from_log(log_file)
    df_e = pd.DataFrame(data['E']).T
    df_q = pd.DataFrame(data['Q']).T

    # Save data to Excel
    with pd.ExcelWriter(log_file.replace('.log', '.xlsx')) as writer:
        df_e.to_excel(writer, sheet_name='Energies')
        df_q.to_excel(writer, sheet_name='Partition Functions')
        write_dict_to_excel(writer, data['Other'], sheet_name='Other')

    print(f"Dati estratti e salvati in {log_file.replace('.log', '.xlsx')}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python script.py <file.log>")
    else:
        log_file = sys.argv[1]
        main(log_file)